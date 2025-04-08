from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import timedelta
import time
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import cv2
import os 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import pyperclip
import matplotlib.pyplot as plt
import pandas as pd
from selenium.common.exceptions import TimeoutException
from dotenv import load_dotenv


# Configurar opções do Chrome
options = Options()
options.add_argument("--start-maximized")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)

# Executar em modo headless
# options.add_argument("--headless=new")

# Desativar o gerenciador de senhas e a oferta de salvar senhas
prefs = {
    "credentials_enable_service": False,  # Desativa o serviço de credenciais
    "profile.password_manager_enabled": False  # Desativa o gerenciador de senhas
}
options.add_experimental_option("prefs", prefs)

# Inicializar navegador com as opções
servico = Service(ChromeDriverManager().install())
navegador = webdriver.Chrome(service=servico, options=options)

load_dotenv()

def apagar_Campo(driver, elemento_id, tempo_espera=30):
    """
    Espera até que o elemento identificado pelo ID seja clicável, clica, limpa o campo e define "0,00".

    :param driver: Instância do WebDriver.
    :param elemento_id: ID do elemento a ser manipulado.
    :param tempo_espera: Tempo máximo de espera em segundos (padrão: 30 segundos).
    """
    try:
        elemento = WebDriverWait(driver, tempo_espera).until(
            EC.element_to_be_clickable((By.ID, elemento_id))
        )
        driver.execute_script(f"document.getElementById('{elemento_id}').click();")
        print("Elemento clicado com sucesso.")

        # Limpando o campo via JavaScript
        driver.execute_script(f"document.getElementById('{elemento_id}').value = '';")
        print("Campo limpo com sucesso.")

        # Definindo o valor como "0,00"
        driver.execute_script(f"document.getElementById('{elemento_id}').value = '0,00';")
        print('Valor "0,00" definido com sucesso.')
    except Exception as e:
        print(f"Erro ao clicar, limpar ou definir o valor do elemento: {e}")

def digitar_entrada_com_TAB(driver, texto, tab_count=0):
    driver.switch_to.active_element.send_keys(texto)
    for _ in range(tab_count):
        driver.switch_to.active_element.send_keys(Keys.TAB)
    time.sleep(1)

def esperar_e_clicar_simples(driver, elemento_id, tempo_espera=30):
    """
    Espera até que o elemento identificado pelo ID seja clicável e, então, realiza um duplo clique.

    :param driver: Instância do WebDriver.
    :param elemento_id: ID do elemento a ser clicado.
    :param tempo_espera: Tempo máximo de espera em segundos (padrão: 30 segundos).
    """
    try:
        # Espera até que o elemento seja clicável
        WebDriverWait(driver, tempo_espera).until(
            EC.element_to_be_clickable((By.ID, elemento_id))
        )

        # Localiza o elemento pelo ID
        elemento = driver.find_element(By.ID, elemento_id)

        # Realiza o duplo clique usando ActionChains
        action = ActionChains(driver)
        action.double_click(elemento).perform()  # Duplo clique
        print('Duplo clique realizado com sucesso.')
    except Exception as e:
        print(f"Erro ao realizar o duplo clique: {e}")

def digitar_entrada(driver, texto, tab_count=0):
    driver.switch_to.active_element.send_keys(texto)
    time.sleep(1)

def esperar_imagem_aparecer(driver, imagem_alvo, timeout=80):
    try:
        if not os.path.exists(imagem_alvo):
            raise FileNotFoundError(f"Imagem alvo não encontrada: {imagem_alvo}")
        
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        
        tempo_espera = 0
        intervalo = 2
        
        while True:
            screenshot = driver.get_screenshot_as_png()
            screen_array = np.frombuffer(screenshot, dtype=np.uint8)
            screen_image = cv2.imdecode(screen_array, cv2.IMREAD_COLOR)
            screen_gray = cv2.cvtColor(screen_image, cv2.COLOR_BGR2GRAY)

            template = cv2.imread(imagem_alvo, cv2.IMREAD_COLOR)
            if template is None:
                raise ValueError(f"Erro ao carregar a imagem do template: {imagem_alvo}")

            template_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
            result = cv2.matchTemplate(screen_gray, template_gray, cv2.TM_CCOEFF_NORMED)
            threshold = 0.7  # Reduzido para melhorar detecção
            _, max_val, _, max_loc = cv2.minMaxLoc(result)

            if max_val >= threshold:
                print(f"Imagem encontrada: {imagem_alvo}")
                return True
            else:
                print("Imagem ainda não visível, aguardando...")
                time.sleep(intervalo)
                tempo_espera += intervalo
                
                if tempo_espera >= timeout:
                    print("Tempo limite atingido. Imagem não encontrada.")
                    return False
    except Exception as e:
        print(f"Erro ao detectar a imagem: {e}")
        return False

def detectar_e_clicar_imagem(driver, imagem_alvo, timeout=80):
    try:
        if not os.path.exists(imagem_alvo):
            raise FileNotFoundError(f"Imagem alvo não encontrada: {imagem_alvo}")

        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        
        tempo_espera = 0
        intervalo = 2
        while tempo_espera < timeout:
            screenshot = driver.get_screenshot_as_png()
            screen_array = np.frombuffer(screenshot, dtype=np.uint8)
            screen_image = cv2.imdecode(screen_array, cv2.IMREAD_COLOR)
            screen_gray = cv2.cvtColor(screen_image, cv2.COLOR_BGR2GRAY)

            template = cv2.imread(imagem_alvo, cv2.IMREAD_COLOR)
            if template is None:
                raise ValueError(f"Erro ao carregar a imagem do template: {imagem_alvo}")

            template_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
            result = cv2.matchTemplate(screen_gray, template_gray, cv2.TM_CCOEFF_NORMED)
            threshold = 0.7  # Reduzido para melhorar detecção
            min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)

            if max_val >= threshold:
                template_h, template_w = template_gray.shape[:2]
                click_x = max_loc[0] + template_w // 2
                click_y = max_loc[1] + template_h // 2

                screen_width = driver.execute_script("return window.innerWidth;")
                screen_height = driver.execute_script("return window.innerHeight;")
                
                click_x = min(max(click_x, 0), screen_width - 1)
                click_y = min(max(click_y, 0), screen_height - 1)
                
                print(f"Tentando clicar em X: {click_x}, Y: {click_y}, Tela: {screen_width}x{screen_height}")
                time.sleep(1)
                webdriver.ActionChains(driver).move_by_offset(click_x, click_y).click().perform()
                return
            else:
                print("Imagem ainda não visível, aguardando...")
                time.sleep(intervalo)
                tempo_espera += intervalo
        
        print("Imagem não encontrada após tempo limite.")
    except Exception as e:
        print(f"Erro ao detectar ou clicar na imagem: {e}")

def Clique_Ousado(driver, imagem_alvo, timeout=80):
    try:
        if not os.path.exists(imagem_alvo):
            print(f"Imagem alvo não encontrada: {imagem_alvo}")
            return
        
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        tempo_espera = 0
        intervalo = 2
        while tempo_espera < timeout:
            screenshot = driver.get_screenshot_as_png()
            screen_array = np.frombuffer(screenshot, dtype=np.uint8)
            screen_image = cv2.imdecode(screen_array, cv2.IMREAD_COLOR)
            screen_gray = cv2.cvtColor(screen_image, cv2.COLOR_BGR2GRAY)

            template = cv2.imread(imagem_alvo, cv2.IMREAD_COLOR)
            if template is None:
                print(f"Erro ao carregar a imagem do template: {imagem_alvo}")
                return
            
            template_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
            result = cv2.matchTemplate(screen_gray, template_gray, cv2.TM_CCOEFF_NORMED)
            
            _, max_val, _, max_loc = cv2.minMaxLoc(result)
            
            print(f"Precisão da imagem encontrada: {max_val:.2f}")

            # Salvar a imagem detectada para análise
            #cv2.imwrite("imagem_detectada.png", screen_image)
            print("Imagem detectada salva como 'imagem_detectada.png'")

            if max_val >= 0.90:  # Ajuste de precisão para aceitar imagens com 98% de correspondência
                template_h, template_w = template_gray.shape[:2]
                click_x = max_loc[0] + template_w // 2
                click_y = max_loc[1] + template_h // 2

                cv2.rectangle(screen_image, (max_loc[0], max_loc[1]), 
                              (max_loc[0] + template_w, max_loc[1] + template_h), 
                              (0, 255, 0), 2)
                #cv2.imwrite("clicado.png", screen_image)
                print("Print da área clicada salvo como 'clicado.png'.")

                scroll_x, scroll_y = driver.execute_script("return [window.scrollX, window.scrollY];")

                # Posição real do clique considerando a rolagem
                page_x = click_x + scroll_x
                page_y = click_y + scroll_y

                print(f"Clicando exatamente em X: {page_x}, Y: {page_y} (Considerando scroll X: {scroll_x}, Y: {scroll_y})")

                # Rolagem até a posição do clique para garantir visibilidade
                driver.execute_script(f"window.scrollTo({page_x - 50}, {page_y - 50});")
                time.sleep(1)

                # Simulação de clique real no ponto exato
                driver.execute_script(f"document.elementFromPoint({click_x}, {click_y}).click();")
                return  
            
            print("Imagem não encontrada com 90% de precisão, aguardando...")
            time.sleep(intervalo)
            tempo_espera += intervalo
        
        print("Imagem não encontrada após tempo limite.")
    except Exception as e:
        print(f"Erro ao detectar ou clicar na imagem: {e}")

def Clique_Ousado_Duas_Vezes(driver, imagem_alvo, timeout=80):
    try:
        if not os.path.exists(imagem_alvo):
            print(f"Imagem alvo não encontrada: {imagem_alvo}")
            return
        
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        tempo_espera = 0
        intervalo = 2
        while tempo_espera < timeout:
            screenshot = driver.get_screenshot_as_png()
            screen_array = np.frombuffer(screenshot, dtype=np.uint8)
            screen_image = cv2.imdecode(screen_array, cv2.IMREAD_COLOR)
            screen_gray = cv2.cvtColor(screen_image, cv2.COLOR_BGR2GRAY)

            template = cv2.imread(imagem_alvo, cv2.IMREAD_COLOR)
            if template is None:
                print(f"Erro ao carregar a imagem do template: {imagem_alvo}")
                return
            
            template_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
            result = cv2.matchTemplate(screen_gray, template_gray, cv2.TM_CCOEFF_NORMED)
            
            _, max_val, _, max_loc = cv2.minMaxLoc(result)
            
            print(f"Precisão da imagem encontrada: {max_val:.2f}")

            # Salvar a imagem detectada para análise
            cv2.imwrite("imagem_detectada.png", screen_image)
            print("Imagem detectada salva como 'imagem_detectada.png'")

            if max_val >= 0.90:  # Ajuste de precisão para aceitar imagens com 90% de correspondência
                template_h, template_w = template_gray.shape[:2]
                click_x = max_loc[0] + template_w // 2
                click_y = max_loc[1] + template_h // 2

                cv2.rectangle(screen_image, (max_loc[0], max_loc[1]), 
                              (max_loc[0] + template_w, max_loc[1] + template_h), 
                              (0, 255, 0), 2)
                cv2.imwrite("clicado.png", screen_image)
                print("Print da área clicada salvo como 'clicado.png'.")

                scroll_x, scroll_y = driver.execute_script("return [window.scrollX, window.scrollY];")

                # Posição real do clique considerando a rolagem
                page_x = click_x + scroll_x
                page_y = click_y + scroll_y

                print(f"Clicando exatamente em X: {page_x}, Y: {page_y} (Considerando scroll X: {scroll_x}, Y: {scroll_y})")

                # Rolagem até a posição do clique para garantir visibilidade
                driver.execute_script(f"window.scrollTo({page_x - 50}, {page_y - 50});")
                time.sleep(1)

                # Criando a ação de duplo clique com Selenium
                elemento = driver.execute_script(f"return document.elementFromPoint({click_x}, {click_y});")
                if elemento:
                    action = ActionChains(driver)
                    action.move_to_element(elemento).double_click().perform()
                    print("Duplo clique realizado!")
                    return
                else:
                    print("Elemento não encontrado para clique.")

            print("Imagem não encontrada com 90% de precisão, aguardando...")
            time.sleep(intervalo)
            tempo_espera += intervalo
        
        print("Imagem não encontrada após tempo limite.")
    except Exception as e:
        print(f"Erro ao detectar ou clicar na imagem: {e}")

def esperar_e_clicar(driver, elemento_id, tempo_espera=30, cliques=2):
    """
    Espera até que o elemento identificado pelo ID esteja presente e realiza múltiplos cliques.

    :param driver: Instância do WebDriver.
    :param elemento_id: ID do elemento a ser clicado.
    :param tempo_espera: Tempo máximo de espera em segundos (padrão: 30 segundos).
    :param cliques: Número de cliques a serem realizados (padrão: 2 - duplo clique).
    """
    try:
        # Espera até que o elemento esteja presente no DOM
        elemento = WebDriverWait(driver, tempo_espera).until(
            EC.presence_of_element_located((By.ID, elemento_id))
        )

        # Espera até que o elemento esteja visível e clicável
        WebDriverWait(driver, tempo_espera).until(
            EC.element_to_be_clickable((By.ID, elemento_id))
        )

        # Executa os cliques necessários
        action = ActionChains(driver)
        for _ in range(cliques):
            action.click(elemento)
        action.perform()
        
        print(f'{cliques} clique(s) realizado(s) com sucesso.')
    except Exception as e:
        print(f"Erro ao clicar no elemento '{elemento_id}': {e}")

def inserir_Sem_Espaço(driver, elemento_id, texto, tempo_espera=30):
    """
    Espera até que o elemento identificado pelo ID seja clicável, limpa o campo e insere o texto.
    Se falhar, insere o texto forçadamente via JavaScript.
    Garante que não haverá espaços extras após a inserção.

    :param driver: Instância do WebDriver.
    :param elemento_id: ID do elemento onde o texto será inserido.
    :param texto: Texto a ser inserido no elemento.
    :param tempo_espera: Tempo máximo de espera em segundos (padrão: 30 segundos).
    """
    texto = texto.strip()  # Remover qualquer espaço em branco no início e no final do texto
    
    try:
        # Espera até o elemento ficar clicável
        elemento = WebDriverWait(driver, tempo_espera).until(
            EC.element_to_be_clickable((By.ID, elemento_id))
        )
        
        # Limpar o campo completamente com JavaScript
        driver.execute_script(f"document.getElementById('{elemento_id}').value = '';")
        print(f"Campo {elemento_id} limpo com sucesso.")

        # Agora, insira o texto sem espaços extras
        driver.execute_script(f"document.getElementById('{elemento_id}').value = '{texto}';")
        
        # Garantir que o foco está no campo para a inserção funcionar corretamente
        driver.execute_script(f"document.getElementById('{elemento_id}').focus();")
        
        print(f"Texto '{texto}' inserido com sucesso via JavaScript no elemento {elemento_id}.")
    except Exception as e:
        print(f"Erro ao inserir texto no elemento {elemento_id}: {e}")

        print(f"Erro ao inserir texto no elemento {elemento_id}: {e}")

def inserir_Com_Python(driver, elemento_id, texto, tempo_espera=30):
    texto = str(texto).strip()  # Converter para string antes de strip()

    try:
        elemento = WebDriverWait(driver, tempo_espera).until(
            EC.presence_of_element_located((By.ID, elemento_id))
        )

        # Verificar se o elemento está habilitado e visível
        if not elemento.is_enabled():
            print(f"O elemento {elemento_id} está desabilitado.")
            return
        if not elemento.is_displayed():
            print(f"O elemento {elemento_id} está oculto.")
            return

        elemento.click()  # Clicar para garantir que está ativo
        time.sleep(0.5)  # Pequeno delay para evitar erros

        elemento.send_keys(Keys.CONTROL + "a")  # Selecionar todo o texto
        elemento.send_keys(Keys.BACKSPACE)  # Apagar texto anterior
        time.sleep(0.5)  

        # Inserir o texto caractere por caractere
        for char in texto:
            elemento.send_keys(char)
            time.sleep(0.1)  # Pequeno delay entre caracteres

        print(f"Texto '{texto}' inserido com sucesso no elemento {elemento_id}.")
    
    except TimeoutException:
        print(f"Erro: Tempo limite excedido ao tentar acessar o elemento {elemento_id}.")
    except Exception as e:
        print(f"Erro inesperado ao inserir texto no elemento {elemento_id}: {e}")
        # Tentar via JavaScript como fallback
        driver.execute_script(f"document.getElementById('{elemento_id}').value = '{texto}';")
        print(f"Texto inserido via JavaScript no elemento {elemento_id}.")

def dividir_em_blocos(lista, tamanho=10):
    for i in range(0, len(lista), tamanho):
        yield lista[i:i + tamanho]

#Variáveis
diretorio = r"C:\Users\Guilherme.Silva\Desktop\gimavecore\GIMAVE\Baixa Credenciados"
dirbordero = r"C:\Users\Guilherme.Silva\Desktop\GIMAVE\Credenciados\Borderôs"
arquivo = "BORDERO 01 - Copia.xlsx"
caminho_arquivo = os.path.join(dirbordero, arquivo)
wb = load_workbook(caminho_arquivo)
ws = wb["Reembolso"]
imagem_ok = os.path.join(diretorio, "ok.png")
imagem_inicio = os.path.join(diretorio, "totvs_inicio.png")
imagem_nome = os.path.join(diretorio, "nome.png")
imagem_favorito = os.path.join(diretorio, "favorito.png")
imagem_funcoes_cpg = os.path.join(diretorio, "funcoescpg.png")
imagem_renovacao = os.path.join(diretorio, "msg_renovacao.png")
imagem_outrasacoes = os.path.join(diretorio, "outras_acoes.png")
imagem_bordero = os.path.join(diretorio, "bordero.png")
imagem_bordero2 = os.path.join(diretorio, "bordero2.png")
imagem_antesbordero = os.path.join(diretorio, "antesbordero.png")
imagem_desmarcar = os.path.join(diretorio, "desmarcar_todos.png")
data = "01/04/2025"

# Abrir página
navegador.get("http://an148124.protheus.cloudtotvs.com.br:1703/webapp/")
WebDriverWait(navegador, 30).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
actions = ActionChains(navegador)
time.sleep(5)

detectar_e_clicar_imagem(navegador, imagem_ok)
time.sleep(5)

esperar_imagem_aparecer(navegador, imagem_inicio)
time.sleep(5)

#Colocar login e senha
digitar_entrada_com_TAB(navegador, os.getenv("LOGIN"), 1)
digitar_entrada_com_TAB(navegador, os.getenv("SENHA"), 1)
actions.send_keys(Keys.ENTER).perform()
time.sleep(5)

esperar_imagem_aparecer(navegador, imagem_nome)
time.sleep(1)

for _ in range(2):
    actions.key_down(Keys.SHIFT).send_keys(Keys.TAB).key_up(Keys.SHIFT).perform()
    time.sleep(0.5)
time.sleep(1)

digitar_entrada_com_TAB(navegador, data ,2)
time.sleep(0.5)
actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
actions.key_down(Keys.CONTROL).send_keys('c').key_up(Keys.CONTROL).perform()
nome_digitado = pyperclip.paste()
print(nome_digitado)

if nome_digitado == '06':
    digitar_entrada(navegador, "06")
    for _ in range(4):
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5)
else:
    digitar_entrada(navegador, "06")
    for _ in range(3):
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5)

#Colocar Modulo
digitar_entrada_com_TAB(navegador, "6",5)
actions.send_keys(Keys.ENTER).perform()
time.sleep(5)

# Clicar em Favorito
Clique_Ousado(navegador, imagem_favorito)
time.sleep(3)

#Funções CPG
Clique_Ousado(navegador, imagem_funcoes_cpg)
time.sleep(3)

try:
    elemento = WebDriverWait(navegador, 40).until(
        EC.presence_of_element_located((By.ID, "COMP4522"))
    )
    elemento.click()
    print("COMP4522 encontrado e clicado.")
except TimeoutException:
    print("COMP4522 não encontrado, seguindo com o código.")

time.sleep(15)

#Msg renovação
#esperar_e_clicar_simples(navegador, "COMP4512")
#time.sleep(7)

esperar_imagem_aparecer(navegador, imagem_antesbordero)
time.sleep(2)

#Outras ações
esperar_e_clicar_simples(navegador, "COMP4606")
time.sleep(2)

#Bordero
esperar_e_clicar_simples(navegador, "COMP4614")
time.sleep(2)

#Bordero 2
esperar_e_clicar_simples(navegador, "COMP4615")
time.sleep(4)

#Capturar número do bordero
elemento = WebDriverWait(navegador, 40).until(EC.presence_of_element_located((By.ID, "COMP6018")))
actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
actions.key_down(Keys.CONTROL).send_keys('c').key_up(Keys.CONTROL).perform()
num_bordero = pyperclip.paste()

#Banco
inserir_Sem_Espaço(navegador,"COMP6022", "756",30)
time.sleep(1)

#Agencia
inserir_Sem_Espaço(navegador,"COMP6023" ,"3337",30)
time.sleep(1)

#Conta
inserir_Sem_Espaço(navegador,"COMP6024", "3780624",30)
time.sleep(1)

#Modelo
inserir_Sem_Espaço(navegador,"COMP6030", "02",30)
time.sleep(2)

#Tipo Pagamento
inserir_Sem_Espaço(navegador,"COMP6031" ,"20",30)
time.sleep(1)

#OK
esperar_e_clicar_simples(navegador, "COMP6032")
time.sleep(3)

#Scroll down para localizar o filtro
tcbrowse = navegador.find_element(By.ID, "COMP6003")

for _ in range(10):  # Tenta rolar várias vezes para garantir visibilidade
    tcbrowse.send_keys(Keys.PAGE_DOWN)
    time.sleep(0.5)

#Selecinar o filtro
actions.send_keys(Keys.ENTER).perform()
time.sleep(2)

#Clicar no Confirmar
esperar_e_clicar(navegador,"COMP6008")
time.sleep(15)


#Desmarcar todos os campos selecionados
wa_element = navegador.find_element(By.ID, "COMP6008")
navegador.execute_script("""
    const shadow = arguments[0].shadowRoot;
    const target = shadow.querySelector('th[id="0"]');
    if (target) {
        target.click();
    } else {
        console.error("Elemento <th id='0'> não encontrado.");
    }
""", wa_element)
time.sleep(5)


# Faz a leitura dos IDs de reembolsos DISPONÍVEIS NA PÁGINA
dados_coluna_5 = navegador.execute_script("""
    const shadow = arguments[0].shadowRoot;
    const tabela = shadow.querySelector("table");
    if (!tabela) return [];

    const linhas = tabela.querySelectorAll("tr");
    const valores = [];

    for (let i = 0; i < linhas.length; i++) {
        const colunas = linhas[i].querySelectorAll("td");
        if (colunas.length > 4) {
            valores.push(colunas[4].innerText.trim());
        }
    }
    return valores;
""", wa_element)
time.sleep(5)

# Faz a leitura dos IDs de reembolsos na PLANILHA
df = pd.read_excel(caminho_arquivo, sheet_name="Reembolso", engine="openpyxl")
valores_coluna_f = df.iloc[1:, 5] #F2 em diante
valores_nao_vazios = valores_coluna_f.dropna().astype(int).apply(lambda x: f"{x:09d}")


fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
ids_erp = set(dados_coluna_5)
total_linhas = len(ids_erp)
ids_excel = {
    f"{int(ws.cell(row=row, column=6).value):09d}": row
    for row in range(2, 2 + total_linhas)
    if ws.cell(row=row, column=6).value is not None and ws.cell(row=row, column=6).fill.start_color.rgb != "FFFFFF00"
}

for bloco in dividir_em_blocos(list(ids_erp), 10):
    for id_erp in bloco:
        if id_erp in ids_excel:
            row = ids_excel[id_erp]
            cell = ws.cell(row=row, column=6)
            actions.send_keys(Keys.ENTER).perform()
            print("Selecionado")
            actions.send_keys(Keys.ARROW_DOWN).perform()
            print("Seta para baixo")
            cell.fill = fill_amarelo
            print("Célula colorida!")
        else:
            actions.send_keys(Keys.ARROW_DOWN).perform()
            print("Não localizado, Seta para baixo...")

        time.sleep(4)  # Aguarda entre cada ação

# Salvar alterações
wb.save(caminho_arquivo)





